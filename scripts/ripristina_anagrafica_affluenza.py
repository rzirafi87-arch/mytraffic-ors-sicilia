import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def ripristina_anagrafica():
    file = "MyTraffic_MASTER.xlsx"
    wb = load_workbook(file)
    ws25 = wb["25_Affluenza_Settimanale_Store"]
    ws02 = wb["02_Negozi"]

    # Varianti accettate per ogni colonna
    store_variants = ["Store_ID", "Store ID", "store_id", "store id", "ID Store", "ID negozio", "ID Negozio"]
    brand_variants = ["Brand", "BRAND", "brand", "Marchio"]
    comune_variants = ["Comune", "comune", "Città", "Citta", "City"]
    indirizzo_variants = ["Indirizzo", "indirizzo", "Address", "Indirizzo completo"]

    # prendi intestazioni
    header02 = {str(ws02.cell(1, c).value).strip(): c for c in range(1, ws02.max_column + 1) if ws02.cell(1, c).value}

    def find_col(header, variants):
        for v in variants:
            if v in header:
                return header[v]
        return None

    col_store = find_col(header02, store_variants)
    col_brand = find_col(header02, brand_variants)
    col_comune = find_col(header02, comune_variants)
    col_indirizzo = find_col(header02, indirizzo_variants)

    # Se non trova la colonna Store_ID, usa la prima colonna disponibile
    if not col_store:
        print("Colonna Store_ID non trovata, uso la prima colonna disponibile.")
        col_store = 1

    # copia dati riga per riga
    for r in range(2, ws02.max_row + 1):
        ws25[f"A{r}"] = ws02.cell(r, col_store).value
        if col_brand:
            ws25[f"B{r}"] = ws02.cell(r, col_brand).value
        if col_comune:
            ws25[f"C{r}"] = ws02.cell(r, col_comune).value
        if col_indirizzo:
            ws25[f"D{r}"] = ws02.cell(r, col_indirizzo).value

    wb.save(file)
    print("✅ Anagrafica ripristinata correttamente")

def aggiungi_peso_store():
    file = "MyTraffic_MASTER.xlsx"
    wb = load_workbook(file)
    ws16 = wb["16_Modello_Gravitazionale"]

    # Trova la colonna Store_ID e la prima riga dati
    header = {str(ws16.cell(1, c).value).strip(): c for c in range(1, ws16.max_column + 1) if ws16.cell(1, c).value}
    col_store = header.get("Store_ID")
    if not col_store:
        print("Store_ID non trovato in 16_Modello_Gravitazionale, uso la prima colonna.")
        col_store = 1

    # Aggiungi colonna Peso_store in fondo
    new_col = ws16.max_column + 1
    ws16.cell(1, new_col).value = "Peso_store"

    # Formula XLOOKUP/CERCA.X (adatta per Excel ITA/ENG)
    for r in range(2, ws16.max_row + 1):
        ws16.cell(r, new_col).value = f'=XLOOKUP({get_column_letter(col_store)}{r},' + "'25_Affluenza_Settimanale_Store'!A:A,'25_Affluenza_Settimanale_Store'!E:E)"

    wb.save(file)
    print("✅ Colonna Peso_store aggiunta in 16_Modello_Gravitazionale")

if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "all"
    if mode == "anagrafica":
        ripristina_anagrafica()
    elif mode == "peso":
        aggiungi_peso_store()
    else:
        ripristina_anagrafica()
        aggiungi_peso_store()
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def ripristina_anagrafica():
    file = "MyTraffic_MASTER.xlsx"
    wb = load_workbook(file)
    ws25 = wb["25_Affluenza_Settimanale_Store"]
    ws02 = wb["02_Negozi"]

    # Varianti accettate per ogni colonna
    store_variants = ["Store_ID", "Store ID", "store_id", "store id", "ID Store", "ID negozio", "ID Negozio"]
    brand_variants = ["Brand", "BRAND", "brand", "Marchio"]
    comune_variants = ["Comune", "comune", "Città", "Citta", "City"]
    indirizzo_variants = ["Indirizzo", "indirizzo", "Address", "Indirizzo completo"]

    # prendi intestazioni
    header02 = {str(ws02.cell(1, c).value).strip(): c for c in range(1, ws02.max_column + 1) if ws02.cell(1, c).value}

    def find_col(header, variants):
        for v in variants:
            if v in header:
                return header[v]
        return None

    col_store = find_col(header02, store_variants)
    col_brand = find_col(header02, brand_variants)
    col_comune = find_col(header02, comune_variants)
    col_indirizzo = find_col(header02, indirizzo_variants)

    # Se non trova la colonna Store_ID, usa la prima colonna disponibile
    if not col_store:
        print("Colonna Store_ID non trovata, uso la prima colonna disponibile.")
        col_store = 1

    # copia dati riga per riga
    for r in range(2, ws02.max_row + 1):
        ws25[f"A{r}"] = ws02.cell(r, col_store).value
        if col_brand:
            ws25[f"B{r}"] = ws02.cell(r, col_brand).value
        if col_comune:
            ws25[f"C{r}"] = ws02.cell(r, col_comune).value
        if col_indirizzo:
            ws25[f"D{r}"] = ws02.cell(r, col_indirizzo).value

    wb.save(file)
    print("✅ Anagrafica ripristinata correttamente")

def aggiungi_peso_store():
    file = "MyTraffic_MASTER.xlsx"
    wb = load_workbook(file)
    ws16 = wb["16_Modello_Gravitazionale"]

    # Trova la colonna Store_ID e la prima riga dati
    header = {str(ws16.cell(1, c).value).strip(): c for c in range(1, ws16.max_column + 1) if ws16.cell(1, c).value}
    col_store = header.get("Store_ID")
    if not col_store:
        print("Store_ID non trovato in 16_Modello_Gravitazionale, uso la prima colonna.")
        col_store = 1

    # Aggiungi colonna Peso_store in fondo
    new_col = ws16.max_column + 1
    ws16.cell(1, new_col).value = "Peso_store"

    # Formula XLOOKUP/CERCA.X (adatta per Excel ITA/ENG)
    for r in range(2, ws16.max_row + 1):
        ws16.cell(r, new_col).value = f'=XLOOKUP({get_column_letter(col_store)}{r},' + "'25_Affluenza_Settimanale_Store'!A:A,'25_Affluenza_Settimanale_Store'!E:E)"

    wb.save(file)
    print("✅ Colonna Peso_store aggiunta in 16_Modello_Gravitazionale")

if __name__ == "__main__":
    mode = sys.argv[1] if len(sys.argv) > 1 else "all"
    if mode == "anagrafica":
        ripristina_anagrafica()
    elif mode == "peso":
        aggiungi_peso_store()
    else:
        ripristina_anagrafica()
        aggiungi_peso_store()       