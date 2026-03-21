from openpyxl import load_workbook

FILE = "MyTraffic_MASTER_validato.xlsx"

wb = load_workbook(FILE)

# -----------------------------
# 23_ORS_IMPORT_TEMPLATE resta sorgente import pulita
# -----------------------------

# -----------------------------
# 22_ORS_MATRIX_CALL
# -----------------------------
ws22 = wb["22_ORS_MATRIX_CALL"]
for r in range(4, ws22.max_row + 1):
    ws22[f"U{r}"] = f'=IFERROR(INDEX(\'23_ORS_IMPORT_TEMPLATE\'!$B:$B,MATCH($T{r},\'23_ORS_IMPORT_TEMPLATE\'!$A:$A,0)),"")'
    ws22[f"V{r}"] = f'=IFERROR(INDEX(\'23_ORS_IMPORT_TEMPLATE\'!$C:$C,MATCH($T{r},\'23_ORS_IMPORT_TEMPLATE\'!$A:$A,0)),"")'
    ws22[f"X{r}"] = f'=IF(AND(U{r}<>"",V{r}<>""),"IMPORTED",IF(S{r}="SI","READY","SKIP"))'

# -----------------------------
# 06_Store_Competitor
# -----------------------------
ws06 = wb["06_Store_Competitor"]
for r in range(4, ws06.max_row + 1):
    ws06[f"Q{r}"]  = f'=IFERROR(INDEX(\'03_Competitor\'!$M:$M,MATCH($F{r},\'03_Competitor\'!$A:$A,0)),0)'
    ws06[f"V{r}"]  = f'=IFERROR(INDEX(\'22_ORS_MATRIX_CALL\'!$U:$U,MATCH($S{r},\'22_ORS_MATRIX_CALL\'!$T:$T,0)),"")'
    ws06[f"W{r}"]  = f'=IFERROR(INDEX(\'22_ORS_MATRIX_CALL\'!$V:$V,MATCH($S{r},\'22_ORS_MATRIX_CALL\'!$T:$T,0)),"")'
    ws06[f"AF{r}"] = f'=IFERROR(INDEX(\'03_Competitor\'!$R:$R,MATCH($F{r},\'03_Competitor\'!$A:$A,0)),"")'
    ws06[f"AG{r}"] = f'=IFERROR(INDEX(\'03_Competitor\'!$S:$S,MATCH($F{r},\'03_Competitor\'!$A:$A,0)),"")'
    ws06[f"AI{r}"] = f'=IFERROR(Q{r},IFERROR(INDEX(\'03_Competitor\'!$M:$M,MATCH(F{r},\'03_Competitor\'!$A:$A,0)),0))'

# -----------------------------
# 16_Modello_Gravitazionale
# -----------------------------
ws16 = wb["16_Modello_Gravitazionale"]
for r in range(7, ws16.max_row + 1):
    ws16[f"I{r}"] = f'=IFERROR(INDEX(\'02_Negozi\'!$H:$H,MATCH(C{r},\'02_Negozi\'!$A:$A,0)),0)'

# -----------------------------
# 02_Negozi
# -----------------------------
ws02 = wb["02_Negozi"]
for r in range(4, ws02.max_row + 1):
    ws02[f"V{r}"] = f'=IFERROR(INDEX(\'11_Bacino_Popolazione\'!$H:$H,MATCH($A{r},\'11_Bacino_Popolazione\'!$A:$A,0)),0)'

# -----------------------------
# 15_Trade_Area
# -----------------------------
ws15 = wb["15_Trade_Area"]
for r in range(4, ws15.max_row + 1):
    ws15[f"F{r}"] = f'=IFERROR(INDEX(\'11_Bacino_Popolazione\'!$B:$B,MATCH(A{r},\'11_Bacino_Popolazione\'!$A:$A,0)),0)'
    ws15[f"G{r}"] = f'=IFERROR(INDEX(\'11_Bacino_Popolazione\'!$D:$D,MATCH(A{r},\'11_Bacino_Popolazione\'!$A:$A,0)),0)'
    ws15[f"H{r}"] = f'=IFERROR(INDEX(\'11_Bacino_Popolazione\'!$H:$H,MATCH(A{r},\'11_Bacino_Popolazione\'!$A:$A,0)),0)'
    ws15[f"J{r}"] = f'=IFERROR(INDEX(\'11_Bacino_Popolazione\'!$T:$T,MATCH(A{r},\'11_Bacino_Popolazione\'!$A:$A,0)),0)'

# -----------------------------
# 18_Ranking_Pro
# -----------------------------
ws18 = wb["18_Ranking_Pro"]
for r in range(4, ws18.max_row + 1):
    ws18[f"J{r}"] = (
        f'=IFERROR(INDEX(\'16_Modello_Gravitazionale\'!$C:$C,'
        f'MATCH(MAXIFS(\'16_Modello_Gravitazionale\'!$P:$P,\'16_Modello_Gravitazionale\'!$A:$A,A{r},\'16_Modello_Gravitazionale\'!$B:$B,B{r}),'
        f'\'16_Modello_Gravitazionale\'!$P:$P,0)),"")'
    )

# forza ricalcolo all'apertura
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(FILE)
print("Patch formule completata.")
