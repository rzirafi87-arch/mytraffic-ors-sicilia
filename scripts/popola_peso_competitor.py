from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
wb = load_workbook(FILE)
ws_comp = wb["03_Competitor"]
ws_sett = wb["01_Impostazioni"]

# Costruisci mappa brand -> peso da 01_Impostazioni
brand_to_peso = {}
for row in ws_sett.iter_rows(min_row=7, max_row=20, min_col=10, max_col=11):
    brand = row[0].value
    peso = row[1].value
    if brand and peso is not None:
        brand_to_peso[str(brand).strip().upper()] = float(peso)

# Trova colonne

# Trova colonne in modo robusto
header = [str(cell.value).strip().lower().replace('_','') if cell.value else '' for cell in ws_comp[3]]
def find_col(name):
    name = name.strip().lower().replace('_','')
    for idx, h in enumerate(header):
        if name in h:
            return idx + 1
    raise ValueError(f"Colonna '{name}' non trovata")

col_brand = find_col("brand")
col_peso = find_col("pesoc")

# Popola Peso_c in base al brand
for r in range(4, ws_comp.max_row + 1):
    brand = ws_comp.cell(row=r, column=col_brand).value
    if brand:
        peso = brand_to_peso.get(str(brand).strip().upper(), 0)
        ws_comp.cell(row=r, column=col_peso, value=peso)
    else:
        ws_comp.cell(row=r, column=col_peso, value=0)

wb.save(FILE)
print("Colonna Peso_c popolata in base alle regole del foglio 01_Impostazioni.")
