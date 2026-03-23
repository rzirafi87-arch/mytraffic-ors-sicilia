import shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

FILE = "MyTraffic_MASTER.xlsx"
BACKUP = "MyTraffic_MASTER_PRE_FASE2.xlsx"

shutil.copy2(FILE, BACKUP)
print(f"Backup creato: {BACKUP}")

wb = load_workbook(FILE)

def norm(x):
    return str(x).strip() if x is not None else ""

def header_map(ws):
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip():
            out[str(v).strip()] = c
    return out

def find_header(ws, candidates):
    h = header_map(ws)
    low = {k.lower(): v for k, v in h.items()}
    for cand in candidates:
        if cand.lower() in low:
            return low[cand.lower()]
    for cand in candidates:
        for k, v in low.items():
            if cand.lower() in k:
                return v
    return None

def clear_rows(ws, start_row):
    if ws.max_row >= start_row:
        for r in range(start_row, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).value = None

def copy_template_row(ws, src_row, dst_row):
    for c in range(1, ws.max_column + 1):
        src = ws.cell(src_row, c)
        dst = ws.cell(dst_row, c)
        if src.data_type == "f" or (isinstance(src.value, str) and str(src.value).startswith("=")):
            try:
                dst.value = Translator(src.value, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                dst.value = src.value
        else:
            dst.value = src.value
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format

def write_if_col(ws, row, col_idx, value):
    if col_idx:
        ws.cell(row, col_idx).value = value

# LETTURA 02_Negozi
ws02 = wb["02_Negozi"]
h02 = header_map(ws02)
c_storeid   = find_header(ws02, ["StoreID", "Store_ID"])
c_comune    = find_header(ws02, ["Comune"])
c_prov      = find_header(ws02, ["Provincia"])
c_brand     = find_header(ws02, ["Brand_Rete", "Brand", "Store"])
c_ind       = find_header(ws02, ["Indirizzo"])
c_lat       = find_header(ws02, ["Lat"])
c_lon       = find_header(ws02, ["Lon"])
c_formato   = find_header(ws02, ["Formato_PV", "Formato"])

stores = []
for r in range(2, ws02.max_row + 1):
    sid = norm(ws02.cell(r, c_storeid).value if c_storeid else "")
    if not sid:
        continue
    stores.append({
        "StoreID": sid,
        "Comune": norm(ws02.cell(r, c_comune).value if c_comune else ""),
        "Provincia": norm(ws02.cell(r, c_prov).value if c_prov else ""),
        "Brand": norm(ws02.cell(r, c_brand).value if c_brand else ""),
        "Indirizzo": norm(ws02.cell(r, c_ind).value if c_ind else ""),
        "Lat": ws02.cell(r, c_lat).value if c_lat else None,
        "Lon": ws02.cell(r, c_lon).value if c_lon else None,
        "Formato": norm(ws02.cell(r, c_formato).value if c_formato else ""),
    })

# LETTURA 03_Competitor
ws03 = wb["03_Competitor"]
h03 = header_map(ws03)
c_compid    = find_header(ws03, ["Competitor_ID", "Competitor"])
c_ccomune   = find_header(ws03, ["Comune"])
c_cprov     = find_header(ws03, ["Provincia"])
c_cbrand    = find_header(ws03, ["Brand_modello", "Brand"])
c_cind      = find_header(ws03, ["Indirizzo"])
c_clat      = find_header(ws03, ["Lat"])
c_clon      = find_header(ws03, ["Lon"])
c_cpeso     = find_header(ws03, ["Peso_competitor"])
c_cdir      = find_header(ws03, ["Competitor_diretto"])
c_clivello  = find_header(ws03, ["Livello_competitor"])

competitors = []
for r in range(2, ws03.max_row + 1):
    cid = norm(ws03.cell(r, c_compid).value if c_compid else "")
    comune = norm(ws03.cell(r, c_ccomune).value if c_ccomune else "")
    if not cid or not comune:
        continue
    competitors.append({
        "Competitor_ID": cid,
        "Comune": comune,
        "Provincia": norm(ws03.cell(r, c_cprov).value if c_cprov else ""),
        "Brand": norm(ws03.cell(r, c_cbrand).value if c_cbrand else ""),
        "Indirizzo": norm(ws03.cell(r, c_cind).value if c_cind else ""),
        "Lat": ws03.cell(r, c_clat).value if c_clat else None,
        "Lon": ws03.cell(r, c_clon).value if c_clon else None,
        "Peso": ws03.cell(r, c_cpeso).value if c_cpeso else None,
        "Diretto": norm(ws03.cell(r, c_cdir).value if c_cdir else ""),
        "Livello": norm(ws03.cell(r, c_clivello).value if c_clivello else ""),
    })

pairs = []
for s in stores:
    for c in competitors:
        pairs.append((s, c))

print(f"Store letti: {len(stores)}")
print(f"Competitor letti: {len(competitors)}")
print(f"Coppie attese: {len(pairs)}")

# RIGENERA 22_ORS_MATRIX_CALL
ws22 = wb["22_ORS_MATRIX_CALL"]
clear_rows(ws22, 2)
needed_last_row = 1 + len(pairs)
for r in range(2, needed_last_row + 1):
    copy_template_row(ws22, 2, r)
h22 = header_map(ws22)
col22 = {
    "Store_Prov": find_header(ws22, ["Store_Prov"]),
    "Competitor": find_header(ws22, ["Competitor"]),
    "Brand": find_header(ws22, ["Brand"]),
    "Comp_Comune": find_header(ws22, ["Comp_Comune"]),
    "Comp_Pro": find_header(ws22, ["Comp_Pro"]),
    "Store_Lat": find_header(ws22, ["Store_Lat"]),
    "Store_Lon": find_header(ws22, ["Store_Lon"]),
    "Comp_La": find_header(ws22, ["Comp_La", "Comp_Lat"]),
    "Comp_Lo": find_header(ws22, ["Comp_Lo", "Comp_Lon"]),
}
r = 2
for s, c in pairs:
    write_if_col(ws22, r, col22["Store_Prov"], s["Provincia"])
    write_if_col(ws22, r, col22["Competitor"], c["Competitor_ID"])
    write_if_col(ws22, r, col22["Brand"], c["Brand"])
    write_if_col(ws22, r, col22["Comp_Comune"], c["Comune"])
    write_if_col(ws22, r, col22["Comp_Pro"], c["Provincia"])
    write_if_col(ws22, r, col22["Store_Lat"], s["Lat"])
    write_if_col(ws22, r, col22["Store_Lon"], s["Lon"])
    write_if_col(ws22, r, col22["Comp_La"], c["Lat"])
    write_if_col(ws22, r, col22["Comp_Lo"], c["Lon"])
    r += 1
print("22_ORS_MATRIX_CALL rigenerato")

# RIGENERA 06_Store_Competitor
ws06 = wb["06_Store_Competitor"]
clear_rows(ws06, 2)
for r in range(2, needed_last_row + 1):
    copy_template_row(ws06, 2, r)
h06 = header_map(ws06)
col06 = {
    "StoreID": find_header(ws06, ["StoreID", "Store_ID"]),
    "Store": find_header(ws06, ["Store"]),
    "Comune": find_header(ws06, ["Comune", "Store_Comune"]),
    "Provincia": find_header(ws06, ["Provincia", "Store_Prov"]),
    "Brand": find_header(ws06, ["Brand", "Brand_Rete"]),
    "Competitor": find_header(ws06, ["Competitor", "Competitor_ID"]),
    "Comp_Comune": find_header(ws06, ["Comp_Comune"]),
    "Comp_Pro": find_header(ws06, ["Comp_Pro"]),
    "Store_Lat": find_header(ws06, ["Store_Lat"]),
    "Store_Lon": find_header(ws06, ["Store_Lon"]),
    "Comp_La": find_header(ws06, ["Comp_La", "Comp_Lat"]),
    "Comp_Lo": find_header(ws06, ["Comp_Lo", "Comp_Lon"]),
    "Indirizzo": find_header(ws06, ["Indirizzo"]),
    "Legacy": find_header(ws06, ["Legacy_key", "Legacy", "Match_legacy", "Store_comp"]),
    "Peso": find_header(ws06, ["Peso_competitor"]),
    "Competitor_diretto": find_header(ws06, ["Competitor_diretto"]),
    "Livello": find_header(ws06, ["Livello_competitor"]),
}
r = 2
for s, c in pairs:
    write_if_col(ws06, r, col06["StoreID"], s["StoreID"])
    write_if_col(ws06, r, col06["Store"], s["Brand"])
    write_if_col(ws06, r, col06["Comune"], s["Comune"])
    write_if_col(ws06, r, col06["Provincia"], s["Provincia"])
    write_if_col(ws06, r, col06["Brand"], c["Brand"])
    write_if_col(ws06, r, col06["Competitor"], c["Competitor_ID"])
    write_if_col(ws06, r, col06["Comp_Comune"], c["Comune"])
    write_if_col(ws06, r, col06["Comp_Pro"], c["Provincia"])
    write_if_col(ws06, r, col06["Store_Lat"], s["Lat"])
    write_if_col(ws06, r, col06["Store_Lon"], s["Lon"])
    write_if_col(ws06, r, col06["Comp_La"], c["Lat"])
    write_if_col(ws06, r, col06["Comp_Lo"], c["Lon"])
    write_if_col(ws06, r, col06["Indirizzo"], c["Indirizzo"])
    write_if_col(ws06, r, col06["Legacy"], f'{s["StoreID"]}_{c["Competitor_ID"]}')
    write_if_col(ws06, r, col06["Peso"], c["Peso"])
    write_if_col(ws06, r, col06["Competitor_diretto"], c["Diretto"])
    write_if_col(ws06, r, col06["Livello"], c["Livello"])
    r += 1
print("06_Store_Competitor rigenerato")

wb.save(FILE)
print(f"\nFASE 2 completata e salvata su: {FILE}")
print("Apri il file, fai Dati > Aggiorna tutto, salva una volta.")
