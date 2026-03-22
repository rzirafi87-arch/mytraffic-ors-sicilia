from openpyxl import load_workbook

xlsx = "MyTraffic_MASTER.xlsx"
sheet_name = "24_Copertura_Competitor"

wb = load_workbook(xlsx)
ws = wb[sheet_name]

for row in range(2, 15):
    ws[f"D{row}"].value = '=IF(OR(B{0}="",B{0}=0),"",B{0}-C{0})'.format(row)
    ws[f"E{row}"].value = '=IF(OR(B{0}="",B{0}=0),"",C{0}/B{0})'.format(row)
    ws[f"F{row}"].value = '=IF(B{0}="","N.D.",IF(C{0}=0,"ASSENTE",IF(C{0}<B{0},"PARZIALE","OK")))'.format(row)

wb.save(xlsx)
print("Formule aggiornate in 24_Copertura_Competitor")
