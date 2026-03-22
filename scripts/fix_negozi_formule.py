from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
wb = load_workbook(FILE)
ws = wb["02_Negozi"]

# Trova l'ultima riga con dati
max_row = ws.max_row

# Colonne di interesse
col_pressione = 'S'  # Pressione_competitiva_totale
col_indice = 'T'     # Pressione_Indice
col_quota = 'U'      # Popolaz_Quota

# Formula per Pressione_competitiva_totale (esempio: conta competitor per comune)
for r in range(2, max_row + 1):
    ws[f"{col_pressione}{r}"] = f"=CONTA.SE('03_Competitor'!C:C;B{r})"

# Formula per Pressione_Indice (esempio: rapporto pressione/popolazione quota)
for r in range(2, max_row + 1):
    ws[f"{col_indice}{r}"] = f"=SE({col_quota}{r}=0;0;{col_pressione}{r}/{col_quota}{r})"

# Formula per Popolaz_Quota (esempio: cerca popolazione per comune)
for r in range(2, max_row + 1):
    ws[f"{col_quota}{r}"] = f"=SE.ERRORE(INDICE('11_Bacino_Popolazione'!D:D;CONFRONTA(B{r};'11_Bacino_Popolazione'!B:B;0));0)"

wb.save(FILE)
print("Formule aggiornate su 02_Negozi (S, T, U)")
