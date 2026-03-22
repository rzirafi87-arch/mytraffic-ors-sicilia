from openpyxl import load_workbook
from collections import defaultdict
import csv
import re
from pathlib import Path

WORKBOOK = "MyTraffic_MASTER.xlsx"
OUTDIR = Path("output")
OUTDIR.mkdir(exist_ok=True)

wb = load_workbook(WORKBOOK, data_only=False)

sheet_summary = []
dependencies = []
formula_rows = []

sheet_ref_pattern = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_]+))!")

for ws in wb.worksheets:
    max_row = ws.max_row
    max_col = ws.max_column
    formula_count = 0
    refs_found = defaultdict(int)

    headers = []
    if max_row >= 1:
        for c in range(1, max_col + 1):
            headers.append(ws.cell(1, c).value)

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(r, c).value
            if isinstance(val, str) and val.startswith("="):
                formula_count += 1
                formula_rows.append([
                    ws.title,
                    ws.cell(r, c).coordinate,
                    val
                ])
                for m in sheet_ref_pattern.finditer(val):
                    ref_sheet = m.group(1) or m.group(2)
                    if ref_sheet and ref_sheet in wb.sheetnames and ref_sheet != ws.title:
                        refs_found[ref_sheet] += 1

    sheet_summary.append([
        ws.title,
        max_row,
        max_col,
        formula_count,
        " | ".join([str(h) for h in headers[:15] if h is not None])
    ])

    for ref_sheet, count in refs_found.items():
        dependencies.append([ws.title, ref_sheet, count])

with open(OUTDIR / "workbook_sheet_summary.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(["sheet_name", "max_row", "max_col", "formula_count", "header_preview"])
    writer.writerows(sheet_summary)

with open(OUTDIR / "workbook_dependencies.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(["from_sheet", "to_sheet", "reference_count"])
    writer.writerows(dependencies)


with open(OUTDIR / "workbook_formulas.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(["sheet_name", "cell", "formula"])
    writer.writerows(formula_rows)

md = []
md.append("# MyTraffic Workbook Audit")
md.append("")
md.append("## Sheet summary")
md.append("")
for row in sheet_summary:
    md.append(f"- **{row[0]}** — righe: {row[1]}, colonne: {row[2]}, formule: {row[3]}")
md.append("")
md.append("## Dependencies")
md.append("")
for row in dependencies:
    md.append(f"- **{row[0]}** -> **{row[1]}** ({row[2]} riferimenti)")
md.append("")

with open(OUTDIR / "workbook_audit.md", "w", encoding="utf-8") as f:
    f.write("\n".join(md))

print("OK - creati:")
print("output/workbook_sheet_summary.csv")
print("output/workbook_dependencies.csv")
print("output/workbook_formulas.csv")
print("output/workbook_audit.md")
