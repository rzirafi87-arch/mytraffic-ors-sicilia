import shutil
from copy import copy
from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
BACKUP = "MyTraffic_MASTER_PRE_FASE1.xlsx"

shutil.copy2(FILE, BACKUP)
print(f"Backup creato: {BACKUP}")

wb = load_workbook(FILE)

def find_col(ws, header_name):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None and str(v).strip() == header_name:
            return c
    return None

def copy_row_style(ws, src_row, dst_row, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for c in range(1, max_col + 1):
        s = ws.cell(src_row, c)
        d = ws.cell(dst_row, c)
        if s.has_style:
            d._style = copy(s._style)
        if s.number_format:
            d.number_format = s.number_format
        if s.font:
            d.font = copy(s.font)
        if s.fill:
            d.fill = copy(s.fill)
        if s.border:
            d.border = copy(s.border)
        if s.alignment:
            d.alignment = copy(s.alignment)
        if s.protection:
            d.protection = copy(s.protection)

def replace_formula_text(v):
    if not isinstance(v, str) or not v.startswith("="):
        return v
    new_v = v
    # Sostituzioni riferimenti colonne
    new_v = new_v.replace("'03_Competitor'!$K:$K", "'03_Competitor'!$S:$S")
    new_v = new_v.replace("'03_Competitor'!$P:$P", "'03_Competitor'!$S:$S")
    new_v = new_v.replace("'03_Competitor'!K:K", "'03_Competitor'!S:S")
    new_v = new_v.replace("'03_Competitor'!P:P", "'03_Competitor'!S:S")
    new_v = new_v.replace("03_Competitor'!$K:$K", "03_Competitor'!$S:$S")
    new_v = new_v.replace("03_Competitor'!$P:$P", "03_Competitor'!$S:$S")
    new_v = new_v.replace("03_Competitor'!K:K", "03_Competitor'!S:S")
    new_v = new_v.replace("03_Competitor'!P:P", "03_Competitor'!S:S")
    new_v = new_v.replace("'03_Competitor'!$A$2:$A$817", "'03_Competitor'!$A$2:$A$930")
    new_v = new_v.replace("'03_Competitor'!$A$2:$A$816", "'03_Competitor'!$A$2:$A$930")
    new_v = new_v.replace("'03_Competitor'!$A$2:$A$929", "'03_Competitor'!$A$2:$A$930")
    # Forza separatore punto e virgola per compatibilità italiana
    if "," in new_v:
        new_v = new_v.replace(",", ";")
    return new_v

def fix_formulas_in_sheet(ws):
    changed = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            nv = replace_formula_text(v)
            if nv != v:
                cell.value = nv
                changed += 1
    return changed

# 1) 03_Competitor
ws = wb["03_Competitor"]
col_id = find_col(ws, "Competitor_ID")
col_comune = find_col(ws, "Comune")
last_real = 1
for r in range(2, ws.max_row + 1):
    v_id = ws.cell(r, col_id).value if col_id else None
    v_com = ws.cell(r, col_comune).value if col_comune else None
    if (v_id is not None and str(v_id).strip() != "") or (v_com is not None and str(v_com).strip() != ""):
        last_real = r
if last_real < 930:
    last_real = 930
if ws.max_row > last_real:
    ws.delete_rows(last_real + 1, ws.max_row - last_real)
print(f"03_Competitor pulito fino a riga {last_real}")
headers_03 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
for c in range(1, ws.max_column + 1):
    if ws.cell(1, c).value == "Priorita_verifica" and c != 22:
        if c != 22:
            ws.cell(1, c).value = "Priorita_verifica_aux"

# 2) 24_Copertura_Competitor
ws = wb["24_Copertura_Competitor"]
if ws.max_row > 14:
    for r in range(15, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).value = None
headers = ["Brand","PV_ufficiali_Sicilia","PV_nel_file","Gap","Copertura","Stato","Note"]
for i, h in enumerate(headers, start=1):
    ws.cell(1, i).value = h
for r in range(2, 15):
    ws[f"D{r}"] = f'=SE(O(B{r}="";B{r}=0);"";B{r}-C{r})'
    ws[f"E{r}"] = f'=SE(O(B{r}="";B{r}=0);"";C{r}/B{r})'
    ws[f"F{r}"] = f'=SE(B{r}="";"DA COMPLETARE";SE(C{r}=0;"ASSENTE";SE(C{r}<B{r};"PARZIALE";"OK")))'
print("24_Copertura_Competitor ripulito")

# 3) 25_Affluenza_Settimanale_Store
ws = wb["25_Affluenza_Settimanale_Store"]
if ws.max_row > 25:
    ws.delete_rows(26, ws.max_row - 25)
print("25_Affluenza_Settimanale_Store ripulito")

# 4) Correzione formule fogli chiave
sheets_to_fix = [
    "04_Comuni",
    "08_Ranking_Comuni",
    "18_Ranking_Pro",
    "13_Pressione_Provincia",
    "10_Dashboard",
    "11_Bacino_Popolazione",
    "07_Traffico_Store",
    "05_Comune_Rete",
    "09_Shortlist",
]
for name in sheets_to_fix:
    if name in wb.sheetnames:
        n = fix_formulas_in_sheet(wb[name])
        print(f"{name}: formule corrette = {n}")

# 5) Riparazioni mirate Dashboard
if "10_Dashboard" in wb.sheetnames:
    ws = wb["10_Dashboard"]
    ws["A1"] = "KPI"
    ws["B1"] = "Valore"
    ws["B3"] = "=CONTA.VALORI('03_Competitor'!A:A)-1"
    ws["B4"] = "=CONTA.SE('03_Competitor'!S:S;\"SI\")"
    if "04_Comuni" in wb.sheetnames:
        ws["B5"] = "=CONTA.VALORI('04_Comuni'!A:A)-1"
    ws["B10"] = "=CONTA.SE('03_Competitor'!Y:Y;\"ALTA\")"
    ws["B11"] = "=CONTA.SE('03_Competitor'!K:K;\"MANCANTE\")"
    ws["B12"] = "=CONTA.PIU.SE('03_Competitor'!G:G;\"\";'03_Competitor'!H:H;\"\")"
    if "06_Store_Competitor" in wb.sheetnames:
        ws["B20"] = "=CONTA.SE('06_Store_Competitor'!Z:Z;\"<>\")"
print("10_Dashboard riparato nei KPI base")

# 6) Foglio 03_Competitor: ricrea formule helper K/L/X/Y
ws = wb["03_Competitor"]
col_ind = find_col(ws, "Indirizzo")
col_com = find_col(ws, "Comune")
col_pro = find_col(ws, "Provincia")
col_lat = find_col(ws, "Lat")
col_lon = find_col(ws, "Lon")
ws["K1"] = "Check_indirizzo"
ws["L1"] = "Query"
ws["X1"] = "Indicatore_record"
ws["Y1"] = "Priorita_verifica"
for r in range(2, ws.max_row + 1):
    if col_ind:
        letter_ind = "F"
    else:
        letter_ind = None
    if letter_ind:
        ws[f"K{r}"] = f'=SE(O(F{r}="";F{r}="NO_ADDR");"MANCANTE";"OK")'
        ws[f"L{r}"] = f'=SE(F{r}="NO_ADDR";"NO_ADDR";F{r})'
    else:
        ws[f"K{r}"] = "MANCANTE"
        ws[f"L{r}"] = "NO_ADDR"
    if col_lat and col_lon:
        ws[f"X{r}"] = f'=SE(E(G{r}<>"";H{r}<>"";K{r}="OK");"READY";SE(K{r}="OK";"PARZIALE";"CHECK"))'
    else:
        ws[f"X{r}"] = f'=SE(K{r}="OK";"PARZIALE";"CHECK")'
    ws[f"Y{r}"] = f'=SE(X{r}="CHECK";"ALTA";SE(X{r}="PARZIALE";"MEDIA";"BASSA"))'
print("03_Competitor helper ripristinati")

wb.save(FILE)
print(f"\nFASE 1 completata e salvata su: {FILE}")
print("Apri il file, fai Dati > Aggiorna tutto, salva una volta.")
