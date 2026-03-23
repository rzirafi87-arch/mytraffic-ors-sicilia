from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "03_Competitor"

wb = load_workbook(FILE)
ws = wb[SHEET]


ws["K1"] = "Check_indirizzo"

for r in range(2, ws.max_row + 1):
    ws[f"K{r}"] = f'=IF(OR(F{r}="",F{r}="NO_ADDR"),"MANCANTE","OK")'

wb.save(FILE)
print("OK - colonna K aggiornata")
