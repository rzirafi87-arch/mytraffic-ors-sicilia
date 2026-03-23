from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "03_Competitor"

wb = load_workbook(FILE)
ws = wb[SHEET]

ws["L1"] = "Query_indirizzo"

for r in range(2, ws.max_row + 1):
    ws[f"L{r}"] = f'=IF(F{r}="",B{r}&" "&C{r}&" Sicilia Italia",F{r}&" "&B{r}&" "&C{r}&" Sicilia Italia")'

wb.save(FILE)
print("OK - colonna L aggiornata con formula compatibile Excel italiano")
