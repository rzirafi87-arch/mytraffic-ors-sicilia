import pandas as pd
from openpyxl import load_workbook

FILE = "MyTraffic_MASTER_validato.xlsx"
ORS = "output/output_ors.csv"

df = pd.read_csv(ORS)
df = df[df["status"] == "OK"].copy()
df["match_key"] = df["store_id"].astype(str) + "_" + df["competitor_id"].astype(str)
df = df[["match_key","distance_km","duration_min"]].drop_duplicates()

wb = load_workbook(FILE)
ws = wb["23_ORS_IMPORT_TEMPLATE"]
ws.delete_rows(1, ws.max_row)

ws["A1"] = "match_key"
ws["B1"] = "distance_km"
ws["C1"] = "duration_min"

for i, row in enumerate(df.itertuples(index=False), start=2):
    ws[f"A{i}"] = row.match_key
    ws[f"B{i}"] = float(row.distance_km)
    ws[f"C{i}"] = float(row.duration_min)

try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(FILE)
print(f"Import ORS completato: {len(df)} righe OK")
