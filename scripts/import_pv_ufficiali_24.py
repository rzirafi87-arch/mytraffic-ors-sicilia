from openpyxl import load_workbook
import csv

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "24_Copertura_Competitor"
CSV_IN = "output/copertura_competitor_da_completare.csv"

wb = load_workbook(FILE)
ws = wb[SHEET]

data = {}
with open(CSV_IN, newline="", encoding="utf-8-sig") as f:
    reader = csv.DictReader(f)
    for row in reader:
        brand = (row["Brand"] or "").strip().upper()
        pv = (row["PV_ufficiali_Sicilia"] or "").strip()
        note = (row["Note"] or "").strip()
        data[brand] = (pv, note)

for r in range(2, 15):
    brand = str(ws[f"A{r}"].value).strip().upper()
    if brand in data:
        pv, note = data[brand]
        ws[f"B{r}"] = pv
        ws[f"G{r}"] = note

wb.save(FILE)
print("OK - valori ufficiali reimportati nel foglio 24_Copertura_Competitor")
