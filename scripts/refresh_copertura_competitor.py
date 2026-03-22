from openpyxl import load_workbook
from collections import Counter

WORKBOOK = "MyTraffic_MASTER.xlsx"
wb = load_workbook(WORKBOOK)

if "03_Competitor" not in wb.sheetnames or "24_Copertura_Competitor" not in wb.sheetnames:
    raise Exception("Manca 03_Competitor o 24_Copertura_Competitor")

ws3 = wb["03_Competitor"]
ws24 = wb["24_Copertura_Competitor"]

# leggi intestazioni 03
h3 = {ws3.cell(1, c).value: c for c in range(1, ws3.max_column + 1)}
brand_col = h3.get("Brand") or h3.get("Brand_modello") or 4

counts = Counter()
for r in range(2, ws3.max_row + 1):
    brand = ws3.cell(r, brand_col).value
    if brand:
        counts[str(brand).strip().upper()] += 1

# prepara intestazioni 24 se mancanti
headers24 = [
    "Brand",
    "PV_ufficiali_Sicilia",
    "PV_nel_file",
    "Gap",
    "Copertura",
    "Stato",
    "Note",
]
for i, h in enumerate(headers24, start=1):
    ws24.cell(1, i).value = h

# scrivi brand e pv nel file
brands_sorted = sorted(counts.keys())
for idx, brand in enumerate(brands_sorted, start=2):
    ws24.cell(idx, 1).value = brand
    ws24.cell(idx, 3).value = counts[brand]
    ws24.cell(idx, 4).value = f'=IF(OR(B{idx}="",C{idx}=""),"",B{idx}-C{idx})'
    ws24.cell(idx, 5).value = f'=IF(OR(B{idx}="",B{idx}=0),"",C{idx}/B{idx})'
    ws24.cell(idx, 6).value = f'=IF(B{idx}="","DA COMPLETARE",IF(C{idx}>=B{idx},"OK","PARZIALE"))'

wb.save(WORKBOOK)
print("OK - foglio 24 aggiornato con PV_nel_file e formule")
