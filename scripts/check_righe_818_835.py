from openpyxl import load_workbook

wb = load_workbook("MyTraffic_MASTER.xlsx", data_only=False)
ws = wb["03_Competitor"]

print("\n=== CHECK RIGHE 818-835 ===")
for r in range(818, 836):
    print(
        r,
        "| D=", ws[f"D{r}"].value,
        "| K=", ws[f"K{r}"].value,
        "| L=", ws[f"L{r}"].value,
        "| X=", ws[f"X{r}"].value,
        "| Y=", ws[f"Y{r}"].value
    )
