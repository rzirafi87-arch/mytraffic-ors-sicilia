from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
wb = load_workbook(FILE)
ws = wb["02_Negozi"]

max_row = ws.max_row

# Colonne di interesse
col_pressione = 'S'  # Pressione_competitiva_totale
col_indice = 'T'     # Pressione_Indice
col_quota = 'U'      # Popolaz_Quota

# Formule in inglese (Excel converte automaticamente in italiano all'apertura)
for r in range(2, max_row + 1):
    ws[f"{col_pressione}{r}"].value = f"=COUNTIF('03_Competitor'!C:C,B{r})"
    ws[f"{col_indice}{r}"].value = f"=IF({col_quota}{r}=0,0,{col_pressione}{r}/{col_quota}{r})"
    ws[f"{col_quota}{r}"].value = f"=IFERROR(INDEX('11_Bacino_Popolazione'!D:D,MATCH(B{r},'11_Bacino_Popolazione'!B:B,0)),0)"

wb.save(FILE)
print("Formule aggiornate in inglese su 02_Negozi (S, T, U) - nessuna corruzione prevista.")
