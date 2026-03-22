from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
wb = load_workbook(FILE)

# -----------------------------
# Foglio 24_Copertura_Competitor
# -----------------------------
ws24 = wb["24_Copertura_Competitor"]

brands = [
    "DECO","LIDL","CONAD","SUPERCONVENIENTE","PAGHI POCO",
    "ARD","MD","EUROSPIN","SISA","CRAI","COOP","IL CENTESIMO","ALTRO"
]

# headers
headers24 = ["Brand","PV_ufficiali_Sicilia","PV_nel_file","Gap","Copertura","Stato","Note"]
for col, h in enumerate(headers24, start=1):
    ws24.cell(row=1, column=col, value=h)

# data rows
for i, brand in enumerate(brands, start=2):
    ws24[f"A{i}"] = brand
    # B = ufficiali Sicilia -> manuale
    # C = conteggio nel file 03_Competitor
    ws24[f"C{i}"] = f'=COUNTIF(\'03_Competitor\'!$K:$K,A{i})'
    # D = gap
    ws24[f"D{i}"] = f'=IF(OR(B{i}="",B{i}=0),"",B{i}-C{i})'
    # E = copertura
    ws24[f"E{i}"] = f'=IF(OR(B{i}="",B{i}=0),"N.D.",C{i}/B{i})'
    # F = stato
    ws24[f"F{i}"] = (
        f'=IF(E{i}="N.D.","N.D.",'
        f'IF(E{i}>=0.95,"OK",IF(E{i}>=0.8,"QUASI OK","DA COMPLETARE")))'
    )

# -----------------------------
# Foglio 25_Affluenza_Settimanale_Store
# -----------------------------
ws25 = wb["25_Affluenza_Settimanale_Store"]

headers25 = [
    "Store_ID","Brand","Comune","Indirizzo","Fonte","Stato_Raccolta",
    "Lun","Mar","Mer","Gio","Ven","Sab","Dom","Media_settimanale","Indice_affluenza"
]
for col, h in enumerate(headers25, start=1):
    ws25.cell(row=1, column=col, value=h)

# svuota righe dati esistenti
if ws25.max_row > 1:
    ws25.delete_rows(2, ws25.max_row - 1)

# carica dati da 02_Negozi
ws02 = wb["02_Negozi"]
out_row = 2
for r in range(4, ws02.max_row + 1):
    store_id = ws02[f"A{r}"].value
    brand = ws02[f"B{r}"].value
    comune = ws02[f"C{r}"].value

    if store_id:
        ws25[f"A{out_row}"] = f'=IFERROR(\'02_Negozi\'!A{r},"")'
        ws25[f"B{out_row}"] = f'=IFERROR(\'02_Negozi\'!B{r},"")'
        ws25[f"C{out_row}"] = f'=IFERROR(\'02_Negozi\'!C{r},"")'
        ws25[f"D{out_row}"] = ""
        ws25[f"E{out_row}"] = "Google Maps"
        ws25[f"F{out_row}"] = "DA FARE"
        ws25[f"N{out_row}"] = f'=IF(COUNT(G{out_row}:M{out_row})=0,"",AVERAGE(G{out_row}:M{out_row}))'
        ws25[f"O{out_row}"] = (
            f'=IF(N{out_row}="","",'
            f'IF(N{out_row}>=80,"ALTA",IF(N{out_row}>=50,"MEDIA","BASSA")))'
        )
        out_row += 1

try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(FILE)
print("Fogli 24 e 25 popolati correttamente.")
