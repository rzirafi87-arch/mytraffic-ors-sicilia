from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
SHEET_COMP = "03_Competitor"
SHEET_COP = "24_Copertura_Competitor"

brand_target = [
    "DECO",
    "LIDL",
    "CONAD",
    "SUPERCONVENIENTE",
    "PAGHI POCO",
    "ARD",
    "MD",
    "EUROSPIN",
    "SISA",
    "CRAI",
    "COOP",
    "IL CENTESIMO",
    "ALTRO"
]

wb = load_workbook(FILE)
ws_comp = wb[SHEET_COMP]
ws_cop = wb[SHEET_COP]

# trova colonna Brand_modello nel foglio 03_Competitor
headers = {}
for c in range(1, ws_comp.max_column + 1):
    v = ws_comp.cell(1, c).value
    if v is not None and str(v).strip():
        headers[str(v).strip()] = c

if "Brand_modello" not in headers:
    raise ValueError("Colonna 'Brand_modello' non trovata in 03_Competitor")

brand_col = headers["Brand_modello"]

# conteggio brand dal foglio competitor
counts = {}
for r in range(2, ws_comp.max_row + 1):
    v = ws_comp.cell(r, brand_col).value
    b = str(v).strip().upper() if v is not None else ""
    if not b:
        b = "ALTRO"
    counts[b] = counts.get(b, 0) + 1

# pulizia foglio 24 nelle prime righe utili
for r in range(1, 40):
    for c in range(1, 8):
        ws_cop.cell(r, c).value = None

# intestazioni
ws_cop["A1"] = "Brand"
ws_cop["B1"] = "PV_ufficiali_Sicilia"
ws_cop["C1"] = "PV_nel_file"
ws_cop["D1"] = "Gap"
ws_cop["E1"] = "Copertura"
ws_cop["F1"] = "Stato"
ws_cop["G1"] = "Note"

# righe tabella
for i, brand in enumerate(brand_target, start=2):
    ws_cop[f"A{i}"] = brand
    ws_cop[f"B{i}"] = ""
    ws_cop[f"C{i}"] = counts.get(brand, 0)
    ws_cop[f"D{i}"] = f'=IF(OR(B{i}="",B{i}=0),"",B{i}-C{i})'
    ws_cop[f"E{i}"] = f'=IF(OR(B{i}="",B{i}=0),"",C{i}/B{i})'
    ws_cop[f"F{i}"] = f'=IF(B{i}="","DA COMPLETARE",IF(C{i}=0,"ASSENTE",IF(C{i}<B{i},"PARZIALE","OK")))'
    ws_cop[f"G{i}"] = ""

wb.save(FILE)
print("OK - tabella base 24_Copertura_Competitor creata")
print("Brand caricati:")
for b in brand_target:
    print(f"{b}: {counts.get(b, 0)}")
