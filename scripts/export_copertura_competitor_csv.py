from openpyxl import load_workbook
import csv

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "24_Copertura_Competitor"
OUT = "output/copertura_competitor_da_completare.csv"

wb = load_workbook(FILE, data_only=False)
ws = wb[SHEET]

rows = []
for r in range(2, 15):
    rows.append([
        ws[f"A{r}"].value,  # Brand
        ws[f"B{r}"].value,  # PV_ufficiali_Sicilia
        ws[f"C{r}"].value,  # PV_nel_file
        ws[f"D{r}"].value,  # Gap
        ws[f"E{r}"].value,  # Copertura
        ws[f"F{r}"].value,  # Stato
        ws[f"G{r}"].value,  # Note
    ])

with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["Brand","PV_ufficiali_Sicilia","PV_nel_file","Gap","Copertura","Stato","Note"])
    w.writerows(rows)

print(f"Creato: {OUT}")
