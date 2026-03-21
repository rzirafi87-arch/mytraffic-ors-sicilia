import pandas as pd
from openpyxl import load_workbook

WORKBOOK = "MyTraffic_MASTER_validato.xlsx"
ORS_CSV = "output/output_ors.csv"

# 1) carica output ORS e pulisce
df = pd.read_csv(ORS_CSV)
df = df[df["status"] == "OK"].copy()
df["match_key"] = df["store_id"].astype(str) + "_" + df["competitor_id"].astype(str)
df = df[["match_key", "distance_km", "duration_min"]].drop_duplicates()

# 2) apre workbook
wb = load_workbook(WORKBOOK)

# 3) riscrive foglio 23_ORS_IMPORT_TEMPLATE
ws23 = wb["23_ORS_IMPORT_TEMPLATE"]
ws23.delete_rows(1, ws23.max_row)

ws23["A1"] = "match_key"
ws23["B1"] = "distance_km"
ws23["C1"] = "duration_min"

for i, row in enumerate(df.itertuples(index=False), start=2):
    ws23[f"A{i}"] = row.match_key
    ws23[f"B{i}"] = float(row.distance_km)
    ws23[f"C{i}"] = float(row.duration_min)

# 4) corregge formule in 22_ORS_MATRIX_CALL
ws22 = wb["22_ORS_MATRIX_CALL"]
for r in range(4, ws22.max_row + 1):
    ws22[f"U{r}"] = f'=IFERROR(XLOOKUP($T{r},\'23_ORS_IMPORT_TEMPLATE\'!$A:$A,\'23_ORS_IMPORT_TEMPLATE\'!$B:$B,"") ,"")'
    ws22[f"V{r}"] = f'=IFERROR(XLOOKUP($T{r},\'23_ORS_IMPORT_TEMPLATE\'!$A:$A,\'23_ORS_IMPORT_TEMPLATE\'!$C:$C,"") ,"")'
    ws22[f"X{r}"] = f'=IF(AND(U{r}<>"",V{r}<>""),"IMPORTED",IF(S{r}="SI","READY","SKIP"))'

# 5) forza ricalcolo all'apertura
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
except Exception:
    pass

wb.save(WORKBOOK)

print("Workbook aggiornato correttamente.")
print(f"Righe ORS OK importate: {len(df)}")
