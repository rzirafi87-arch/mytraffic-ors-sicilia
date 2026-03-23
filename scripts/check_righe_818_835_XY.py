from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "03_Competitor"

wb = load_workbook(FILE, data_only=True)
ws = wb[SHEET]

print("Righe 818-835, colonne G, H, K, X, Y:")
for r in range(818, 836):
    vals = [ws[f"{col}{r}"].value for col in ["G", "H", "K", "X", "Y"]]
    print(f"Riga {r}: {vals}")
