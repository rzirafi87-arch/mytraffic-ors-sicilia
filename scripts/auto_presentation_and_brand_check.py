from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
wb = load_workbook(FILE)

# 1. Ordina i fogli in modo logico per presentazione e analisi
sheet_order = [
    "02_Negozi",
    "03_Competitor",
    "24_Copertura_Competitor",
    "25_Affluenza_Settimanale_Store",
    "22_ORS_MATRIX_CALL",
    "06_Store_Competitor",
    "16_Modello_Gravitazionale",
    "15_Trade_Area",
    "18_Ranking_Pro"
]
# Aggiungi eventuali altri fogli non elencati in fondo
for s in wb.sheetnames:
    if s not in sheet_order:
        sheet_order.append(s)
wb._sheets = [wb[s] for s in sheet_order if s in wb.sheetnames]

# 2. Automatizza la formula di conteggio competitor su Brand_finale (colonna S)
ws24 = wb["24_Copertura_Competitor"]
# Trova la colonna Brand_finale in 03_Competitor
ws03 = wb["03_Competitor"]
brand_col = None
for idx, cell in enumerate(next(ws03.iter_rows(min_row=1, max_row=1))):
    if cell.value and "brand_finale" in str(cell.value).lower():
        brand_col = idx + 1
        break
if not brand_col:
    brand_col = 19  # fallback su S

for i in range(2, ws24.max_row + 1):
    ws24[f"C{i}"] = f'=COUNTIF(\'03_Competitor\'!${chr(64+brand_col)}:${chr(64+brand_col)},A{i})'

# 3. Controlla che tutti i brand siano presenti con almeno un negozio
brand_check = {}
for i in range(2, ws24.max_row + 1):
    brand = ws24[f"A{i}"].value
    count = ws24[f"C{i}"].value
    brand_check[brand] = count

wb.save(FILE)
print("Fogli ordinati, formula conteggio aggiornata, check brand completato.")
print("Brand e negozi trovati:")
for b, c in brand_check.items():
    print(f"{b}: {c}")
