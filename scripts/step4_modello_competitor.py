import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Parametri
excel_path = "MyTraffic_MASTER.xlsx"
sheet_name = "03_Competitor"
missing_dir = "output/negozi_mancanti_per_brand"

# Carica workbook e foglio
wb = load_workbook(excel_path)
ws = wb[sheet_name]

# Trova intestazioni
header = [cell.value for cell in ws[1]]

# Per ogni file di mancanti
for file in os.listdir(missing_dir):
    if not file.endswith(".csv"): continue
    brand = file.replace("_missing.csv", "")
    df = pd.read_csv(os.path.join(missing_dir, file))
    if df.empty: continue
    # Allinea colonne a quelle del foglio Excel
    row_template = {h: "" for h in header}
    # Prova a mappare colonne comuni
    for idx, row in df.iterrows():
        new_row = row_template.copy()
        for col in df.columns:
            if col in new_row:
                new_row[col] = row[col]
        # Imposta brand se manca
        if "Brand" in new_row and not new_row["Brand"]:
            new_row["Brand"] = brand
        # Aggiungi riga al foglio
        ws.append([new_row[h] for h in header])
    print(f"Aggiunti {len(df)} mancanti per {brand}")

# Salva
wb.save(excel_path)
print("Integrazione completata. Tutti i mancanti sono stati aggiunti al foglio 03_Competitor.")
import re
import unicodedata
from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "03_Competitor"

TARGET = [
    "DECO","LIDL","CONAD","SUPERCONVENIENTE","PAGHI POCO",
    "ARD","MD","EUROSPIN","SISA","CRAI","COOP","IL CENTESIMO"
]

def norm(x):
    if x is None:
        return ""
    x = str(x).upper().strip()
    x = unicodedata.normalize("NFKD", x).encode("ascii","ignore").decode("ascii")
    x = re.sub(r"[^A-Z0-9 ]"," ",x)
    x = re.sub(r"\s+"," ",x)
    return x

def brand_map(x):
    x = norm(x)

    if "DECO" in x: return "DECO"
    if "LIDL" in x: return "LIDL"
    if "CONAD" in x: return "CONAD"
    if "EUROSPIN" in x: return "EUROSPIN"
    if "MD" in x: return "MD"
    if "ARD" in x: return "ARD"
    if "SISA" in x: return "SISA"
    if "CRAI" in x: return "CRAI"
    if "COOP" in x: return "COOP"
    if "CENTESIMO" in x: return "IL CENTESIMO"
    if "SUPERCONVENIENTE" in x or "SUPER CONVENIENTE" in x: return "SUPERCONVENIENTE"
    if "PAGHI POCO" in x or "PAGHIPOCO" in x: return "PAGHI POCO"

    return "ALTRO"

wb = load_workbook(FILE)
ws = wb[SHEET]

# intestazioni
headers = {ws.cell(1,c).value:c for c in range(1,ws.max_column+1) if ws.cell(1,c).value}

def col(name_list):
    for n in name_list:
        if n in headers:
            return headers[n]
    return None

c_nome = col(["Nome","Insegna","Denominazione","Brand"])
c_comune = col(["Comune","Citta"])
c_ind = col(["Indirizzo"])
c_lat = col(["Lat"])
c_lon = col(["Lon"])

# colonne target
cols = {
    "Classe_competitor":"Classe_competitor",
    "Famiglia_brand":"Famiglia_brand",
    "Ready_layer":"Ready_layer",
    "Dedupe_key":"Dedupe_key",
    "Livello_competitor":"Livello_competitor",
    "Competitor_diretto":"Competitor_diretto",
    "Brand_modello":"Brand_modello",
    "Flag_brand_chiave":"Flag_brand_chiave",
    "Priorita_verifica":"Priorita_verifica"
}

# crea colonne se mancanti
for k in cols:
    if k not in headers:
        col_n = ws.max_column+1
        ws.cell(1,col_n).value = k
        headers[k]=col_n

# aggiorna indici
idx = {k:headers[k] for k in cols}

for r in range(2, ws.max_row+1):

    nome = ws.cell(r,c_nome).value if c_nome else ""
    comune = ws.cell(r,c_comune).value if c_comune else ""
    indirizzo = ws.cell(r,c_ind).value if c_ind else ""
    lat = ws.cell(r,c_lat).value if c_lat else ""
    lon = ws.cell(r,c_lon).value if c_lon else ""

    brand = brand_map(nome)

    flag = 1 if brand in TARGET else 0

    # valori
    classe = "KEY_COMPETITOR" if flag else "ALTRO"
    famiglia = brand
    diretto = "SI" if flag else "NO"
    livello = "DIRETTO" if flag else "ALTRO"

    dedupe = norm(f"{brand} {comune} {indirizzo}")

    if lat and lon:
        ready = "READY"
    elif indirizzo:
        ready = "PARZIALE"
    else:
        ready = "CHECK"

    if ready == "CHECK":
        priorita = "ALTA"
    elif ready == "PARZIALE":
        priorita = "MEDIA"
    else:
        priorita = "BASSA"

    # scrittura
    ws.cell(r, idx["Classe_competitor"]).value = classe
    ws.cell(r, idx["Famiglia_brand"]).value = famiglia
    ws.cell(r, idx["Ready_layer"]).value = ready
    ws.cell(r, idx["Dedupe_key"]).value = dedupe
    ws.cell(r, idx["Livello_competitor"]).value = livello
    ws.cell(r, idx["Competitor_diretto"]).value = diretto
    ws.cell(r, idx["Brand_modello"]).value = brand
    ws.cell(r, idx["Flag_brand_chiave"]).value = flag
    ws.cell(r, idx["Priorita_verifica"]).value = priorita

wb.save(FILE)

print("🔥 STEP 4 COMPLETATO - MODELLO COMPETITOR ATTIVO")
