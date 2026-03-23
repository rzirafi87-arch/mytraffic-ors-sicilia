from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "03_Competitor"

wb = load_workbook(FILE)
ws = wb[SHEET]


ws["X1"] = "Indicatore_record"
ws["Y1"] = "Priorita_verifica"

for r in range(2, ws.max_row + 1):
    ws[f"X{r}"] = f'=IF(AND(G{r}<>"",H{r}<>"",K{r}="OK"),"READY",IF(K{r}="OK","PARZIALE","CHECK"))'
    ws[f"Y{r}"] = f'=IF(X{r}="CHECK","ALTA",IF(X{r}="PARZIALE","MEDIA","BASSA"))'

wb.save(FILE)
print("OK - colonne X e Y aggiornate")
