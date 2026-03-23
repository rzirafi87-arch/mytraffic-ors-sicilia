from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "03_Competitor"

wb = load_workbook(FILE)
ws = wb[SHEET]

# Mappa intestazioni
headers = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(1, c).value
    if v is not None and str(v).strip():
        headers[str(v).strip()] = c

def first_existing(candidates):
    for x in candidates:
        if x in headers:
            return headers[x]
    return None

col_nome = first_existing(["Nome", "nome", "Insegna", "insegna", "Denominazione", "Brand", "brand"])
col_ind = first_existing(["Indirizzo", "indirizzo", "Address", "address"])
col_com = first_existing(["Comune", "comune", "Citta", "citta", "City", "city"])
col_pro = first_existing(["Provincia", "provincia", "Prov", "prov"])
col_lat = first_existing(["Lat", "lat", "Latitude", "latitude"])
col_lon = first_existing(["Lon", "lon", "Lng", "lng", "Longitude", "longitude"])

# colonne target fisse richieste da te
# K = 11, L = 12, X = 24, Y = 25
ws["K1"] = "Check_indirizzo"
ws["L1"] = "Query_indirizzo"
ws["X1"] = "Indicatore_record"
ws["Y1"] = "Priorita_verifica"

for r in range(2, ws.max_row + 1):
    # 1) se colonna indirizzo esiste ed è vuota, la ricostruisco
    if col_ind:
        val_ind = ws.cell(r, col_ind).value
        if val_ind is None or str(val_ind).strip() == "":
            via = ""
            comune = ws.cell(r, col_com).value if col_com else ""
            provincia = ws.cell(r, col_pro).value if col_pro else ""
            nome = ws.cell(r, col_nome).value if col_nome else ""
            pezzi = [str(x).strip() for x in [nome, comune, provincia, "Sicilia", "Italia"] if x not in (None, "")]
            ricostruito = ", ".join(pezzi)
            ws.cell(r, col_ind).value = ricostruito

    ind_letter = get_column_letter(col_ind) if col_ind else None
    com_letter = get_column_letter(col_com) if col_com else None
    pro_letter = get_column_letter(col_pro) if col_pro else None
    lat_letter = get_column_letter(col_lat) if col_lat else None
    lon_letter = get_column_letter(col_lon) if col_lon else None

    # 2) K = controllo indirizzo
    if ind_letter:
        ws[f"K{r}"] = f'=IF(TRIM({ind_letter}{r})="","MANCANTE","OK")'
    else:
        ws[f"K{r}"] = 'COLONNA INDIRIZZO NON TROVATA'

    # 3) L = query / indirizzo completo
    parts = []
    if ind_letter: parts.append(f'{ind_letter}{r}')
    if com_letter: parts.append(f'{com_letter}{r}')
    if pro_letter: parts.append(f'{pro_letter}{r}')
    if parts:
        join_formula = ",".join(parts)
        ws[f"L{r}"] = f'=TEXTJOIN(", ",TRUE,{join_formula},"Sicilia","Italia")'
    else:
        ws[f"L{r}"] = 'QUERY NON COSTRUIBILE'

    # 4) X = indicatore record
    if ind_letter and lat_letter and lon_letter:
        ws[f"X{r}"] = f'=IF(AND(TRIM({ind_letter}{r})<>"",{lat_letter}{r}<>"",{lon_letter}{r}<>""),"READY","CHECK")'
    elif ind_letter:
        ws[f"X{r}"] = f'=IF(TRIM({ind_letter}{r})<>"","PARZIALE","CHECK")'
    else:
        ws[f"X{r}"] = 'CHECK'

    # 5) Y = priorità da X
    ws[f"Y{r}"] = f'=IF(X{r}="CHECK","ALTA",IF(X{r}="PARZIALE","MEDIA","BASSA"))'

wb.save(FILE)
print("OK - aggiornati indirizzi mancanti e formule in K, L, X, Y")
