import re
import unicodedata
import csv
from openpyxl import load_workbook

XLSX = "MyTraffic_MASTER.xlsx"
SHEET_COMP = "03_Competitor"
SHEET_COP = "24_Copertura_Competitor"

TARGET_BRANDS = [
    "DECO", "LIDL", "CONAD", "SUPERCONVENIENTE", "PAGHI POCO",
    "ARD", "MD", "EUROSPIN", "SISA", "CRAI", "COOP", "IL CENTESIMO"
]

HELPER_COLS = [
    "Classe_competitor", "Famiglia_brand", "Ready_layer", "Dedupe_key",
    "Livello_competitor", "Competitor_diretto", "Brand_modello",
    "Flag_brand_chiave", "Priorita_verifica"
]

def norm_text(x):
    if x is None:
        return ""
    s = str(x).strip().upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.replace("&", " E ")
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def canon_brand(raw):
    s = norm_text(raw)
    if not s:
        return "ALTRO"
    if "DECO" in s:
        return "DECO"
    if "LIDL" in s:
        return "LIDL"
    if "CONAD" in s:
        return "CONAD"
    if "SUPERCONVENIENTE" in s or "SUPER CONVENIENTE" in s:
        return "SUPERCONVENIENTE"
    if "PAGHI POCO" in s or "PAGHIPOCO" in s or "PAGHI POCO FRATELLO" in s or "PAGHI POCO FRATELLI" in s:
        return "PAGHI POCO"
    if s == "ARD" or s.startswith("ARD "):
        return "ARD"
    if s == "MD" or s.startswith("MD ") or "MD DISCOUNT" in s:
        return "MD"
    if "EUROSPIN" in s:
        return "EUROSPIN"
    if "SISA" in s:
        return "SISA"
    if "CRAI" in s:
        return "CRAI"
    if s == "COOP" or s.startswith("COOP "):
        return "COOP"
    if "IL CENTESIMO" in s or "CENTESIMO" in s:
        return "IL CENTESIMO"
    return "ALTRO"

def is_number(v):
    try:
        if v is None or str(v).strip() == "":
            return False
        float(str(v).replace(",", "."))
        return True
    except:
        return False

def safe_slug(*parts):
    txt = " | ".join(norm_text(p) for p in parts if p is not None)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt[:250]

def col_letter(n):
    s = ""
    while n:
        n, rem = divmod(n - 1, 26)
        s = chr(65 + rem) + s
    return s

wb = load_workbook(XLSX)
ws = wb[SHEET_COMP]

headers = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(1, c).value
    if v is not None and str(v).strip():
        headers[str(v).strip()] = c

for col_name in HELPER_COLS:
    if col_name not in headers:
        new_col = ws.max_column + 1
        ws.cell(1, new_col).value = col_name
        headers[col_name] = new_col

def first_existing(candidates):
    for name in candidates:
        if name in headers:
            return headers[name]
    return None

col_brand_raw = first_existing([
    "Brand", "brand", "Insegna", "insegna", "Denominazione",
    "denominazione", "Nome", "nome", "Ragione_sociale", "ragione_sociale"
])
col_comune = first_existing(["Comune", "comune", "Citta", "citta", "City", "city"])
col_indirizzo = first_existing(["Indirizzo", "indirizzo", "Address", "address", "Via", "via"])
col_lat = first_existing(["Lat", "lat", "Latitude", "latitude"])
col_lon = first_existing(["Lon", "lon", "Lng", "lng", "Longitude", "longitude"])

c_classe = headers["Classe_competitor"]
c_famiglia = headers["Famiglia_brand"]
c_ready = headers["Ready_layer"]
c_dedupe = headers["Dedupe_key"]
c_livello = headers["Livello_competitor"]
c_diretto = headers["Competitor_diretto"]
c_brand_modello = headers["Brand_modello"]
c_flag = headers["Flag_brand_chiave"]
c_priorita = headers["Priorita_verifica"]

brand_counts = {}

for r in range(2, ws.max_row + 1):
    raw_brand = ws.cell(r, col_brand_raw).value if col_brand_raw else ""
    comune = ws.cell(r, col_comune).value if col_comune else ""
    indirizzo = ws.cell(r, col_indirizzo).value if col_indirizzo else ""
    lat = ws.cell(r, col_lat).value if col_lat else ""
    lon = ws.cell(r, col_lon).value if col_lon else ""

    brand_modello = canon_brand(raw_brand)
    famiglia_brand = brand_modello
    brand_counts[brand_modello] = brand_counts.get(brand_modello, 0) + 1

    flag_brand = 1 if brand_modello in TARGET_BRANDS else 0
    competitor_diretto = "SI" if flag_brand == 1 else "NO"
    livello = "DIRETTO" if flag_brand == 1 else "ALTRO"
    classe = "KEY_COMPETITOR" if flag_brand == 1 else "ALTRO_COMPETITOR"
    dedupe = safe_slug(brand_modello, comune, indirizzo)

    has_geo = is_number(lat) and is_number(lon)
    has_min = brand_modello != "" and comune not in (None, "") and indirizzo not in (None, "")
    ready = "READY" if has_geo and has_min else "CHECK"

    if flag_brand == 1 and not has_geo:
        priorita = "ALTA"
    elif flag_brand == 1:
        priorita = "MEDIA"
    else:
        priorita = "BASSA"

    ws.cell(r, c_classe).value = classe
    ws.cell(r, c_famiglia).value = famiglia_brand
    ws.cell(r, c_ready).value = ready
    ws.cell(r, c_dedupe).value = dedupe
    ws.cell(r, c_livello).value = livello
    ws.cell(r, c_diretto).value = competitor_diretto
    ws.cell(r, c_brand_modello).value = brand_modello
    ws.cell(r, c_flag).value = flag_brand
    ws.cell(r, c_priorita).value = priorita

ws2 = wb[SHEET_COP]
brand_modello_col_letter = col_letter(c_brand_modello)

for r in range(2, 15):
    ws2[f"C{r}"] = f"=COUNTIF('03_Competitor'!${brand_modello_col_letter}:${brand_modello_col_letter},A{r})"
    ws2[f"D{r}"] = f'=IF(OR(B{r}="",B{r}=0),"",B{r}-C{r})'
    ws2[f"E{r}"] = f'=IF(OR(B{r}="",B{r}=0),"",C{r}/B{r})'
    ws2[f"F{r}"] = f'=IF(B{r}="","N.D.",IF(C{r}=0,"ASSENTE",IF(C{r}<B{r},"PARZIALE","OK")))'

wb.save(XLSX)

with open("output/pv_brand_modello.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.writer(f)
    w.writerow(["Brand_modello", "PV_nel_file"])
    for k in sorted(brand_counts.keys()):
        w.writerow([k, brand_counts[k]])

print("OK - popolato 03_Competitor")
print("OK - aggiornato 24_Copertura_Competitor")
print("Creato: output/pv_brand_modello.csv")
for k in sorted(brand_counts.keys()):
    print(f"{k}: {brand_counts[k]}")
