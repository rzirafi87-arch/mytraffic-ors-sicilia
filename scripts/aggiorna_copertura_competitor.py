from openpyxl import load_workbook
import pandas as pd

xlsx = "MyTraffic_MASTER.xlsx"
csv_file = "output/copertura_competitor_base.csv"
sheet_name = "24_Copertura_Competitor"

df = pd.read_csv(csv_file)
wb = load_workbook(xlsx)
ws = wb[sheet_name]

# scrive intestazioni
for c, col in enumerate(df.columns, start=1):
    ws.cell(row=1, column=c, value=col)

# scrive dati
for r, row in enumerate(df.itertuples(index=False), start=2):
    for c, value in enumerate(row, start=1):
        ws.cell(row=r, column=c, value=value)

wb.save(xlsx)
print(f"Aggiornato foglio '{sheet_name}' in {xlsx}")
