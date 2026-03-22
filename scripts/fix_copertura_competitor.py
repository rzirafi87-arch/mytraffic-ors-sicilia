from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
wb = load_workbook(FILE)
ws = wb["24_Copertura_Competitor"]

# Trova intestazione e colonne
header = [str(cell.value).strip().upper() if cell.value else '' for cell in ws[1]]
col_brand = 1  # Si assume che la colonna A sia il brand da cercare
col_conteggio = 3  # Si assume che la colonna C sia quella del conteggio (adatta se necessario)

# Applica la formula COUNTIF su tutte le righe (in inglese, Excel la convertirà)
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=col_conteggio).value = f"=COUNTIF('03_Competitor'!D:D,A{r})"

wb.save(FILE)
print("Formula di conteggio competitor aggiornata su 24_Copertura_Competitor.")
