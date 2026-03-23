from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
SHEET = "03_Competitor"

wb = load_workbook(FILE)
ws = wb[SHEET]

for r in range(818, 836):
    # Esempio: se G e H sono valorizzati, metti "OK" in K
    if ws[f"G{r}"].value not in (None, "") and ws[f"H{r}"].value not in (None, ""):
        ws[f"K{r}"] = "OK"
    else:
        ws[f"K{r}"] = "CHECK"

wb.save(FILE)
print("OK - colonna K aggiornata per le righe 818-835")
