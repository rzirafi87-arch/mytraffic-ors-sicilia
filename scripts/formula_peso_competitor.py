from openpyxl import load_workbook

FILE = "MyTraffic_MASTER.xlsx"
wb = load_workbook(FILE)
ws = wb["03_Competitor"]

# Mappa brand -> peso (dall'immagine)
brand_to_peso = {
    "SUPERCONVENIENTE": 4,
    "PAGHI POCO": 4,
    "CONAD": 3,
    "DECO": 3,
    "COOP": 3,
    "IL CENTESIMO": 3,
    "ARD": 2,
    "LIDL": 2,
    "MD": 2,
    "SISA": 1.5,
    "EUROSPIN": 1.5,
    "CRAI": 1.5,
    "ALTRO": 0
}

# Trova intestazione e colonne

# Trova colonne in modo robusto
header = [str(cell.value).strip().lower().replace('_','') if cell.value else '' for cell in ws[3]]
def find_col(name):
    name = name.strip().lower().replace('_','')
    for idx, h in enumerate(header):
        if name in h:
            return idx + 1
    raise ValueError(f"Colonna '{name}' non trovata")

col_brand = find_col("brand")
col_peso = find_col("pesoc")

# Applica formula in inglese (Excel la convertirà in italiano)
for r in range(4, ws.max_row + 1):
    cell_brand = ws.cell(row=r, column=col_brand)
    formula = (
        '=IF(D{r}="SUPERCONVENIENTE",4,'
        'IF(D{r}="PAGHI POCO",4,'
        'IF(D{r}="CONAD",3,'
        'IF(D{r}="DECO",3,'
        'IF(D{r}="COOP",3,'
        'IF(D{r}="IL CENTESIMO",3,'
        'IF(D{r}="ARD",2,'
        'IF(D{r}="LIDL",2,'
        'IF(D{r}="MD",2,'
        'IF(D{r}="SISA",1.5,'
        'IF(D{r}="EUROSPIN",1.5,'
        'IF(D{r}="CRAI",1.5,0))))))))))))'
    ).replace('{r}', str(r))
    ws.cell(row=r, column=col_peso).value = formula

wb.save(FILE)
print("Formula Peso_c inserita in tutte le righe del foglio 03_Competitor.")
