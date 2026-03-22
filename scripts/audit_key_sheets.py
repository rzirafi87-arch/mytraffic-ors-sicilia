from openpyxl import load_workbook
import csv
from pathlib import Path

WORKBOOK = "MyTraffic_MASTER.xlsx"
OUTDIR = Path("output")
OUTDIR.mkdir(exist_ok=True)

target_sheets = [
    "01_Impostazioni",
    "02_Negozi",
    "03_Competitor",
    "04_Comuni",
    "05_Comune_Rete",
    "06_Store_Competitor",
    "11_Bacino_Popolazione",
    "15_Trade_Area",
    "16_Modello_Gravitazionale",
    "18_Ranking_Pro",
    "22_ORS_MATRIX_CALL",
    "24_Copertura_Competitor",
    "25_Affluenza_Settimanale_Store",
]

wb = load_workbook(WORKBOOK, data_only=False)
rows = []

for s in target_sheets:
    if s not in wb.sheetnames:
        rows.append([s, "MISSING", "", "", ""])
        continue

    ws = wb[s]
    headers = []
    for c in range(1, min(ws.max_column, 25) + 1):
        headers.append(ws.cell(1, c).value)

    sample_formulas = []
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, min(ws.max_column, 40) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                sample_formulas.append(f"{ws.cell(r,c).coordinate}: {v}")
            if len(sample_formulas) >= 8:
                break
        if len(sample_formulas) >= 8:
            break

    rows.append([
        s,
        "OK",
        ws.max_row,
        ws.max_column,
        " | ".join([str(h) for h in headers if h is not None]),
        " || ".join(sample_formulas)
    ])

with open(OUTDIR / "workbook_key_sheets_audit.csv", "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    writer.writerow(["sheet_name", "status", "max_row", "max_col", "headers_preview", "sample_formulas"])
    writer.writerows(rows)

print("OK - creato output/workbook_key_sheets_audit.csv")
